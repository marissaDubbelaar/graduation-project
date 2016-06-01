#!/usr/bin/env python

# Title			: 	createOverviewXlsx.py
# Created by	: 	Marissa Dubbelaar
# Created on	: 	01-03-2016
# Purpose		:	Create an overview showing all of the necessary information about the study and samples.

from openpyxl import load_workbook
from openpyxl import Workbook
import os

class CreateOverviewXlsx():
	def __init__(self, pathway):
		'''
			An overview is made with the use of a pathway and a file name.
			Several functions within this class are called seperately from the defined main.
			The main function only takes care of the creation of the files or obtaining the known accession numbers.
		'''
		# The pathway, name and the extention of the file is given.
		self.pathway = pathway + ".xlsx"
		# Known Accession numbers are saved within self.knownAccessions
		self.knownAccessions = []
		# Known SRA numbers are saved in self.knownSra.
		self.knownSra = []
		# The function self.createXlsx is called.
		self.createXlsx()

	def createXlsx(self):
		'''
			createXlsx is used as the main function. 
			Cheking if the file name exists already or that it should be made.
			The necessary functions with these processes are called.
			In the end the file is saved.
		'''
		# If the file exists already.
		if os.path.isfile(self.pathway):
			# The file is loaded and the sheets "GOAD_studies" and "attributes" are obtained.
			self.overview = load_workbook(self.pathway)
			self.goadStudiesSheet = self.overview.get_sheet_by_name("GOAD_studies")
			self.attrSheet = self.overview.get_sheet_by_name("attributes")
			# All of the known accession numbers are checked that are already known within the file.
			self.checkAvailabilityOfAccessionNr()
		# When the file is unknown
		else:
			# A new workbook is made.
			self.overview = Workbook()
			self.overview.save(self.pathway)
			# The sheet "Sheet" is renamed to "GOAD_studies".
			self.goadStudiesSheet = self.overview.get_sheet_by_name("Sheet")
			self.goadStudiesSheet.title = "GOAD_studies"
			# A sheet "attributes is made"
			self.attrSheet = self.overview.create_sheet()
			self.attrSheet.title = "attributes"
			# A header is added for both the sheet "GOAD_studies" and "attributes"
			self.addHeaderGoadStudySheet()
			self.addHeaderAttributeSheet()
			# All of the infromation is added in the attribute sheet.
			self.addInfoAttributeSheet()
		# The number of the last empty row is obtained.
		self.number = self.obtainEmptyRow()
		# The file is saved.
		self.saveXlsx()

	def writeToXlsx(self, information):
		'''
			The function writeToXlsx is used to write all of the necessary information into the "GOAD_studies" sheet.
		'''
		# The last empty row is obtained
		self.number = self.obtainEmptyRow()
		# If the accession number is not known in self.knownAccessions.
		if information[0] not in self.knownAccessions:
			# The index is defined.
			self.goadStudiesSheet['A'+str(self.number)] = self.number - 1
			# The accession (E-GEOD) number is filled in.
			self.goadStudiesSheet['B'+str(self.number)] = information[0]
			# The author is filled in.
			self.goadStudiesSheet['C'+str(self.number)] = information[1]
			# The year is filled in.
			self.goadStudiesSheet['D'+str(self.number)] = information[2]
			# The title is filled in.
			self.goadStudiesSheet['E'+str(self.number)] = information[3]
			# The abstract is filled in.
			self.goadStudiesSheet['F'+str(self.number)] = information[4]
			# The journal is filled in.
			self.goadStudiesSheet['G'+str(self.number)] = information[5]
			# The organism is filled in.
			self.goadStudiesSheet['H'+str(self.number)] = information[6]
			# The platform is filled in.
			self.goadStudiesSheet['I'+str(self.number)] = information[7]
			# The paired/single end information is filled in.
			self.goadStudiesSheet['J'+str(self.number)] = information[8]
			# The tissue is filled in.
			self.goadStudiesSheet['K'+str(self.number)] = information[9]
			# The cell type is filled in.
			self.goadStudiesSheet['L'+str(self.number)] = information[10]
			# The strain/mutation is filled in.
			self.goadStudiesSheet['M'+str(self.number)] = information[11]
			# The age is filled in.
			self.goadStudiesSheet['N'+str(self.number)] = information[12]
			# The SRA number is filled in.
			self.goadStudiesSheet['O'+str(self.number)] = information[13]
		self.saveXlsx()

	def saveXlsx(self):
		'''
			Saving the overview file.
		'''
		self.overview.save(self.pathway)

	def addHeaderAttributeSheet(self):
		'''
			This function creates the standard header for the attributes sheet.
		'''
		self.attrSheet["A1"] = "name"
		self.attrSheet["B1"] = "entity"
		self.attrSheet["C1"] = "dataType"
		self.attrSheet["D1"] = "idAttribute"
		self.attrSheet["E1"] = "nillable"
		self.attrSheet["F1"] = "description"

	def addInfoAttributeSheet(self):
		'''
			Information from the header of the "GOAD_studies" overview are used to write into the "attributes" sheet.
			Adding the necessary information like the ID, and other information such are the type of information about the attribute.
		'''
		firstRow = True
		for row in self.goadStudiesSheet.iter_rows("A1:O1"):
			iteration = 2
			for item in row:
				if firstRow == True:
					self.attrSheet['A'+str(iteration)] = item.value
					self.attrSheet['B'+str(iteration)] = self.attrSheet.title
					self.attrSheet['E'+str(iteration)] = "FALSE"
					self.attrSheet['F'+str(iteration)] = item.value
					if iteration == 2:
						self.attrSheet['D'+str(iteration)] = "TRUE"
					if item.value == "Title" or item.value == "Abstract" or item.value == "Celltype":
						self.attrSheet['C'+str(iteration)] = "text" 
					elif item.value == "Unique_ID" or item.value == "Year":
						self.attrSheet['C'+str(iteration)] = "int" 
					iteration += 1

	def addHeaderGoadStudySheet(self):
		'''
			The header information for the sheet "GOAD_studies".
		'''
		self.goadStudiesSheet['A1'] = 'Unique_ID'
		self.goadStudiesSheet['B1'] = 'GEOD_NR'
		self.goadStudiesSheet['C1'] = 'Author'
		self.goadStudiesSheet['D1'] = 'Year'
		self.goadStudiesSheet['E1'] = 'Title'
		self.goadStudiesSheet['F1'] = 'Abstract'
		self.goadStudiesSheet['G1'] = 'Journal'
		self.goadStudiesSheet['H1'] = 'Organism'
		self.goadStudiesSheet['I1'] = 'Platform'
		self.goadStudiesSheet['J1'] = 'PE_SE'
		self.goadStudiesSheet['K1'] = 'Tissue'
		self.goadStudiesSheet['L1'] = 'Celltype'
		self.goadStudiesSheet['M1'] = 'Strain'
		self.goadStudiesSheet['N1'] = 'Age'
		self.goadStudiesSheet['O1'] = 'SRA'

	def obtainEmptyRow(self):
		'''
			The function obtainEmptyRow is used to check for the first empty cell in column A.
			The empty row number is returned in the end.
		'''
		colA = "A"
		rowNum = 1
		filledCell = True
		# while filledCell == True.
		while filledCell:
			# If the value within column A + rownumer is empty.
			if self.goadStudiesSheet[colA+str(rowNum)].value== None:
				#  filledCell is changed into False.
				filledCell= False
			else:
				# Otherwise the row number is increased.
				rowNum += 1
		# Returning the number of the row.
		return rowNum

	def checkAvailabilityOfAccessionNr(self):
		'''
			Obtains the found SRA and accession numbers that are already known within the overview.
		'''
		# For each row within the complex overview (without the header).
		for row in self.goadStudiesSheet.iter_rows(row_offset=1):
			# Add the known accession and SRA in the defined list.
			self.knownAccessions.insert(len(self.knownAccessions), str(row[14].value))
			self.knownSra.insert(len(self.knownSra), str(row[1].value))

	def returnKnownSra(self):
		'''
			Returns the list self.knownSra.
		'''
		return self.knownSra

	def returnKnownAcc(self):
		'''
			Returns the list self.knownAccessions.
		'''
		return self.knownAccessions