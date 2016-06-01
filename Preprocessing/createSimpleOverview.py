#!/usr/bin/env python

# Title			: 	createSimpleOVerview.py
# Created by	: 	Marissa Dubbelaar
# Created on	: 	07-03-2016
# Purpose		:	Create a simplefied overview, showing only study information

from openpyxl import load_workbook
from openpyxl import Workbook

import os

class CreateSimpleOverview():
	def __init__(self):
		'''
			The complete overview is used to create a simplefied new xlsx file. 
			This file only shows the studies and their information (information about the samples is left out).
		'''
		# The pathway to the created (complex) overview
		self.usedXlsx = "/Users/marissa/Desktop/GOAD_GliaOnly_Overview_17_3_2016.xlsx"
		# Pathway and name for the nem file.
		self.pathway = "/Users/marissa/Desktop/GOAD_Overview.xlsx"
		# Several empty lists are created.
		self.uniqueEGEOD = []
		self.uniqueStudies = []
		self.knownEGEOD = []
		# The main function (self.createXlsxFile) is called.
		self.createXlsxFile()

	def createXlsxFile(self):
		'''
			The function createXlsxFile is used to obtain a sheet within a xlsx file where the information
			can be written to.
		'''
		# The xlsx file is opened and the sheet "GOAD_studies" is loaded.
		self.oldOverview = load_workbook(self.usedXlsx)
		self.goadStudiesSheet = self.oldOverview.get_sheet_by_name("GOAD_studies")
		# The if-else statement check if the simplefied overview exists already.
		if os.path.isfile(self.pathway):
			# When it already exists new information is added into the first empty row.
			self.newOverview = load_workbook(self.pathway)
			# The sheet with the name 'Overview' is used.
			self.overviewSheet = self.newOverview.get_sheet_by_name("Overview")
			# The known EGEOD numbers within the simplefied overview are obtained.
			self.checkAvailabilityOfEGEOD()
			# The first empty row number is obtained for further use.
			self.number = self.obtainEmptyRow()
		else:
			# If the file is unknown, a new sheet is made. 
			# Also providing the header.
			self.newOverview = Workbook()
			# The sheet with the name 'Sheet' (normal for a new sheet) is used.
			self.overviewSheet = self.newOverview.get_sheet_by_name("Sheet")
			# The name is changed to 'Overview'.
			self.overviewSheet.title = "Overview"
			# A header is written within the sheet.
			self.createHeader()
			# The new overview is saved.
			self.newOverview.save(self.pathway)
		# Studies are obtained from the complex overview.
		self.obtainKnownStudies()
		# Obtained studies are written into the simplefied overview.
		self.writeKnownStudies()
		# The whole file is saved in the end.
		self.newOverview.save(self.pathway)

	def obtainKnownStudies(self):
		'''
			This function check the content of the more complex overview.
			Returning the unique information about studies.
		'''
		# For each row (excluding the header).
		for row in self.goadStudiesSheet.iter_rows(row_offset=1):
			# When the EGEOD number is not known within self.uniqueEGEOD and != None (last line is empty).
			if row[1].value not in self.uniqueEGEOD and row[1].value != None:
				# Information about the study is added.
				self.uniqueStudies.append([row[1].value.encode("utf-8"), row[2].value.encode("utf-8"), row[3].value.encode("utf-8"), row[4].value.encode("utf-8"), row[6].value.encode("utf-8"), row[7].value.encode("utf-8"), row[8].value.encode("utf-8")])
				# The EGEOD number is added into self.uniqueEGEOD.
				self.uniqueEGEOD.append(row[1].value)

	def createHeader(self):
		'''
			The function createHeader is used when the overview file doesn't exists already.
		'''
		self.overviewSheet['A1'] = 'GEOD_NR'
		self.overviewSheet['B1'] = 'Author'
		self.overviewSheet['C1'] = 'Year'
		self.overviewSheet['D1'] = 'Title'
		self.overviewSheet['E1'] = 'Journal'
		self.overviewSheet['F1'] = 'Organism'
		self.overviewSheet['G1'] = 'Platform'

	def writeKnownStudies(self):
		'''
			The number where of the first empty row is obtained.
			Each study within the list self.uniqueStudies that is not known within the
			list of already known studies is written into the simplefied overview.
			Saving the overview in the end.
		'''
		# The first empty row is obtained.
		self.number = self.obtainEmptyRow()
		# For the studies within the found unique studies.
		for study in self.uniqueStudies:
			# When this study is unknown within the simplefied file.
			if study[0] not in self.knownEGEOD:
				# Write all of the information into the file
				self.overviewSheet['A'+str(self.number)] = study[0]
				self.overviewSheet['B'+str(self.number)] = study[1]
				self.overviewSheet['C'+str(self.number)] = study[2]
				self.overviewSheet['D'+str(self.number)] = study[3]
				self.overviewSheet['E'+str(self.number)] = study[4]
				self.overviewSheet['F'+str(self.number)] = study[5]
				self.overviewSheet['G'+str(self.number)] = study[6]
				# Add + 1 to self.number to make sure the next study is written on the line below.
				self.number += 1
		# The overview is saved.
		self.newOverview.save(self.pathway)

	def obtainEmptyRow(self):
		'''
			The function obtainEmptyRow is used to check for the first empty cell in column A.
			The empty row number is returned in the end.
		'''
		colA = "A"
		rowNum = 1
		filledCell = True
		# while filledCell == True
		while filledCell:
			# If the value within column A + rownumer is empty
			if self.overviewSheet[colA + str(rowNum)].value == None:
				#  filledCell is changed into False.
				filledCell= False
			else:
				# Otherwise the row number is increased.
				rowNum += 1
		# Returning the number of the row.
		return rowNum

	def checkAvailabilityOfEGEOD(self):
		'''
			This function checks the already known studies within the simplefied overview.
		'''
		# For each row within the complex overview (without the header)
		for row in self.overviewSheet.iter_rows(row_offset=1):
			# Inset the EGEOD number into self.knownEGEOD.
			self.knownEGEOD.insert(len(self.knownEGEOD), str(row[1].value))

if __name__ == "__main__":
	cso = CreateSimpleOverview()
